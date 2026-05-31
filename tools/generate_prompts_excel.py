import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LLM Prompts Inventory"

# Headers
headers = [
    "File Path",
    "Prompt Name / Area",
    "Lines / Source",
    "Purpose",
    "How to Call (API / Trigger)",
    "When Recalled",
    "Frontend Visible?",
    "Where Shown on Frontend",
    "Stored in MongoDB?",
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_alignment
    c.border = thin_border

# Column widths
ws.column_dimensions["A"].width = 55
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 45
ws.column_dimensions["E"].width = 42
ws.column_dimensions["F"].width = 35
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 40
ws.column_dimensions["I"].width = 16

body_font = Font(size=10)
wrap_alignment = Alignment(wrap_text=True, vertical="top")

data = [
    # ── BACKEND (analysis.service.ts) ──
    [
        "backend/src/analysis/analysis.service.ts",
        "DAILY_PROMPT",
        "Lines 4-29",
        "Review a single day of training. Outputs 3-5 bullet points: session summary, quality, key metrics, weather/terrain, one actionable insight. Edge cases: rest day, short ride, multi-activity clustering.",
        "POST /analysis { type: 'daily', activities: [...] }",
        "User clicks Daily Review button or sends /day command in PaceBotChat",
        "Yes",
        "PaceBotChat (/day command) · Statistics page (Daily Review button → AnalysisModal)",
        "No — ephemeral text returned to caller",
    ],
    [
        "backend/src/analysis/analysis.service.ts",
        "WEEKLY_PROMPT",
        "Lines 31-59",
        "Review a full week (Mon-Sun) as a training block. Covers totals, consistency, intensity distribution, load trend vs prior weeks, recovery signals, one recommendation. Edge cases: incomplete/zero/race/extreme-volume weeks.",
        "POST /analysis { type: 'weekly', activities: [...] }",
        "User clicks Weekly Review button, AI Review button in WeeklyGraph, or sends /week command in PaceBotChat",
        "Yes",
        "PaceBotChat (/week command) · Statistics page (Weekly Review button → AnalysisModal) · WeeklyGraph AI Review button",
        "No — ephemeral text returned to caller",
    ],
    [
        "backend/src/analysis/analysis.service.ts",
        "MONTHLY_PROMPT",
        "Lines 61-91",
        "Review a full month as a macro-cycle phase. Covers totals, volume trend vs prior month, frequency, intensity breakdown, terrain diversity, progression, fatigue, goal alignment, next-month recommendation.",
        "POST /analysis { type: 'monthly', activities: [...] }",
        "User clicks Monthly Review button or sends /month command in PaceBotChat",
        "Yes",
        "PaceBotChat (/month command) · Statistics page (Monthly Review button → AnalysisModal)",
        "No — ephemeral text returned to caller",
    ],
    [
        "backend/src/analysis/analysis.service.ts",
        "CHAT_SYSTEM_PROMPT",
        "Lines 93-120",
        "General Q&A cycling coach persona. Principles: consistency, recovery, adaptation, honesty. Behavior: use FTP zones, explain why, flag overtraining. Communication: bullets, cycling terminology. Response-length rules by question complexity.",
        "POST /analysis { type: 'chat', message: '...', activities: [...] }",
        "User sends a message in PaceBotChat or types a follow-up in AnalysisModal",
        "Yes",
        "PaceBotChat (floating chat widget) · AnalysisModal (follow-up questions)",
        "No — ephemeral text returned to caller",
    ],
    [
        "backend/src/analysis/analysis.service.ts",
        "generateNextWeekPlan prompt",
        "Lines 252-285 (inline)",
        "Generate a 7-day training plan as strict JSON. Provides athlete context (FTP, weight, goal), recent activity summary. Requests { workouts[{dayOfWeek, type, distance, zoneBreakdown, terrain, notes}], coachNotes }. Constraints: 1-2 rest days, 1 long ride, correct intensity distribution, distance in km.",
        "POST /analysis/generate-plan { activities: [...] }",
        "User clicks + Next Week button in WeekScheduleCard",
        "Yes",
        "WeekScheduleCard (+ Next Week button) — plan is saved and displayed",
        "No (by analysis service) — but frontend saves result via POST /training-context/weekly-plan → MongoDB WeeklyPlan",
    ],

    # ── PACKAGES/SPORT-CYCLING (core agent prompts) ──
    [
        "packages/sport-cycling/SOUL.md",
        "soul (cycling coach identity)",
        "1-86",
        "Defines the cycling coaching identity: check fitness/fatigue/form first, consistency, recovery, adaptability, honesty. Power zones only, explain why, flag overtraining, save details to memory. Communication: bullets, cycling terminology, no padding. Cycling-specific review rules.",
        "Injected into system prompt by buildSystemPrompt() in packages/core/src/agent/system-prompt.ts",
        "On every core agent chat() call",
        "No",
        "— Only used by CLI/Telegram coach agent (packages/core)",
        "No — part of system prompt, not stored",
    ],
    [
        "packages/sport-cycling/DAILY_REVIEW.md",
        "Daily Review (core agent variant)",
        "1-25",
        "Same purpose as backend DAILY_PROMPT but used by the core agent for /review commands. 3-5 bullet points. Edge cases: rest day, short ride, multi-activity.",
        "Injected into system prompt via buildSystemPrompt() under ANALYSIS_PROMPTS.daily",
        "On /review command in core agent CLI/Telegram coach",
        "No",
        "— Core agent only",
        "No — ephemeral chat reply",
    ],
    [
        "packages/sport-cycling/WEEKLY_REVIEW.md",
        "Weekly Review (core agent variant)",
        "1-28",
        "Same as backend WEEKLY_PROMPT but for core agent. Weekly totals, consistency, intensity distribution, load trend, recovery, one recommendation.",
        "Injected into system prompt via buildSystemPrompt() under ANALYSIS_PROMPTS.weekly",
        "On /review command in core agent CLI/Telegram coach",
        "No",
        "— Core agent only",
        "No — ephemeral chat reply",
    ],
    [
        "packages/sport-cycling/MONTHLY_REVIEW.md",
        "Monthly Review (core agent variant)",
        "1-30",
        "Same as backend MONTHLY_PROMPT but for core agent. Monthly totals, volume trend, frequency, intensity breakdown, progression, fatigue, goal alignment, next-month recommendation.",
        "Injected into system prompt via buildSystemPrompt() under ANALYSIS_PROMPTS.monthly",
        "On /review command in core agent CLI/Telegram coach",
        "No",
        "— Core agent only",
        "No — ephemeral chat reply",
    ],
    [
        "packages/sport-cycling/WEEKLY_PLAN_INSTRUCTIONS.md",
        "Weekly Plan Instructions",
        "1-116",
        "LLM receives deterministic WeekSkeleton JSON and renders athlete-friendly plan. Specifies JSON output format, zone breakdown convention, workout types, volume constraints (max 10% increase, 3:1 build:recovery), terrain mapping, weather consideration. Prohibitions: no inventing volume, no changing taper timing, no moving key sessions.",
        "Core agent tool build_plan_skeleton → LLM generates rendered week as JSON",
        "When plan is generated/rendered by core agent (CLI/Telegram)",
        "No",
        "— Core agent only (plan data can be viewed via POST /training-context/weekly-plan API)",
        "Yes — saved as MongoDB WeeklyPlan document via API",
    ],
    [
        "packages/sport-cycling/PRE_RACE_PLAN_INSTRUCTIONS.md",
        "Pre-Race Plan Instructions",
        "1-108",
        "LLM receives deterministic TaperSchedule + race info and generates PreRaceWeekPlan documents. Covers week structure by offset (Week -4 Build through Week -1 Race Week), taper rules (volume 70%/50% of peak, maintain intensity, no VO2max in taper, openers 2-3 days before), terrain specificity.",
        "Core agent tool build_taper_week → LLM generates pre-race weeks as JSON",
        "When taper/race prep is generated by core agent (CLI/Telegram)",
        "No",
        "— Core agent only (plan data via GET /training-context/pre-race/:raceId API)",
        "Yes — saved as MongoDB PreRaceWeekPlan documents via API",
    ],
    [
        "packages/sport-cycling/NUTRITION_PLAN_INSTRUCTIONS.md",
        "Nutrition Plan Instructions",
        "1-88",
        "LLM creates DietPlan (multi-day meal plan) and RaceNutrition (race-day fueling) documents. Includes meal timing guidelines table, carb rate 60-90g/h, dual-source carbs, hydration 500-750ml/h, caffeine 3-6mg/kg.",
        "Core agent tool (triggered by athlete chat request in CLI/Telegram)",
        "When athlete requests nutrition plan via core agent chat",
        "No",
        "— Core agent only",
        "Yes — saved as MongoDB DietPlan and RaceNutrition documents via API",
    ],
    [
        "packages/sport-cycling/TRAINING_CONTEXT_SUMMARY.md",
        "Training Context Summary Instructions",
        "1-86",
        "LLM summarizes training into MonthContext and WeekContext documents. MonthContext: single paragraph under 200 chars (headline numbers, key sessions, consistency, phase, trend). WeekContext: under 120 chars. Includes rotation rules (keep last 2 months/weeks).",
        "Core agent tool or scheduled trigger (month/week ISO change)",
        "On month/week change or explicit request in core agent",
        "No",
        "— Core agent only (context data via GET /training-context/months, /training-context/weeks API)",
        "Yes — saved as MongoDB MonthContext and WeekContext documents",
    ],

    # ── PACKAGES/SPORT-CYCLING SKILLS ──
    [
        "packages/sport-cycling/skills/intervals-icu.md",
        "Skill: intervals.icu API reference",
        "1-170",
        "Key metrics (CTL/ATL/TSB/NP/IF/VI), power curve interpretation, athlete type identification, wellness data, calendar workout pushing via intervals_create_workup tool with JSON step definitions.",
        "Injected into system prompt as domain knowledge via persona.skills[]",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt context",
    ],
    [
        "packages/sport-cycling/skills/periodization.md",
        "Skill: Periodization models",
        "1-55",
        "5 periodization models: Linear, Block, Reverse Linear, Polarized (80/20), Pyramidal. Selection logic by experience level/time. Phase allocation with intensity distributions. Build:Recovery ratios (2:1 to 4:1).",
        "Injected into system prompt as domain knowledge via persona.skills[]",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt context",
    ],
    [
        "packages/sport-cycling/skills/race-prep.md",
        "Skill: Race preparation",
        "1-56",
        "Taper duration by race type (century=2wk, criterium=1wk), taper strategy, race-week structure, opener workout design, pre-race nutrition reminders.",
        "Injected into system prompt as domain knowledge via persona.skills[]",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt context",
    ],
    [
        "packages/sport-cycling/skills/recovery.md",
        "Skill: Recovery protocols",
        "1-58",
        "Overtraining warning signs (data & behavioral), when to replace hard sessions, active recovery workout design, deload week design (60-70% volume), sleep/stress impact.",
        "Injected into system prompt as domain knowledge via persona.skills[]",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt context",
    ],
    [
        "packages/sport-cycling/skills/review.md",
        "Skill: Review analysis",
        "1-97",
        "Decoupling interpretation, best-efforts duration ladder, fade pattern recognition, indoor vs outdoor signal differences, show-numbers follow-up table format.",
        "Injected into system prompt as domain knowledge via persona.skills[]",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt context",
    ],
    [
        "packages/sport-cycling/skills/workout-design.md",
        "Skill: Workout design",
        "1-131",
        "Workout type reference (Endurance Z2 through Sprint), structures, cadences, progressive overload patterns, indoor vs outdoor guidance, intervals_create_workout JSON examples.",
        "Injected into system prompt as domain knowledge via persona.skills[]",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt context",
    ],
    [
        "packages/sport-cycling/skills/zone-reference.md",
        "Skill: Power zone reference",
        "1-30",
        "6-zone table (Z1 Active Recovery through Z6 VO2max) with %FTP ranges, RPE, descriptions, HR cross-reference by %max HR.",
        "Injected into system prompt as domain knowledge via persona.skills[]",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt context",
    ],

    # ── PACKAGES/CORE (agent internals) ──
    [
        "packages/core/src/agent/system-prompt.ts",
        "WORKOUT_REVIEW_RULES",
        "Lines 8-91",
        "Instructions for /review command processing. Covers trigger detection, argument parsing, session selection logic, depth tiers, mandatory 3-question framework (Did it go well? / One thing to fix? / What does this mean for next session?), prose-only output, trademark forbidden terms (NP/TSS/IF/CTL/ATL/TSB).",
        "Injected into system prompt by buildSystemPrompt() placed last for recency",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt",
    ],
    [
        "packages/core/src/agent/system-prompt.ts",
        "toolsNote",
        "Lines 102-112",
        "Instructions explaining available Strava tools: strava_fetch_activities for time-sensitive queries, strava_search_history as fallback. Warns against mixing MongoDB vs Strava IDs.",
        "Injected into system prompt by buildSystemPrompt()",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — part of system prompt",
    ],
    [
        "packages/core/src/agent/memory-flush.ts",
        "MEMORY_FLUSH_SYSTEM_PROMPT",
        "Lines 19-22",
        "Extracts athlete info from conversation before summarization. Instructs LLM to use memory_write tool, include ALL current facts for each section, not just new ones.",
        "Memory compaction/flush in core agent loop",
        "When conversation is being compacted to avoid overflow",
        "No",
        "— Core agent only (internal memory management)",
        "No — writes to MEMORY.md on disk, not MongoDB",
    ],
    [
        "packages/core/src/agent/compaction.ts",
        "buildSummarizePrompt / buildDroppedMessagesPrompt",
        "Lines 68-94",
        "Summarizes conversation for overflow handling. Must preserve: athlete profile (FTP, weight), training plan status, recent feedback, decisions, injuries, last topic. Output headings: Athlete Profile, Training Status, Discussion Context, Pending Questions.",
        "Core agent compaction when chat exceeds token limit",
        "When chat history exceeds token limit and needs compaction",
        "No",
        "— Core agent only (internal chat management)",
        "No — stored in chat JSONL on disk as [Previous conversation summary]",
    ],
    [
        "packages/core/src/embeddings/analysis.ts",
        "buildPrompt(ride) — ride analysis",
        "Lines 20-66",
        "Analyze a single ride for Pinecone embeddings. Returns JSON: sessionTitle, coachSummary (3-6 sentences), loadNotes, pacingNotes, softTags. Includes ride data: power zones, pacing structure, fade %, VI, surge count, interval details.",
        "EmbeddingSync.processActivity() → calls buildPrompt() → LLM returns JSON",
        "When new activity is synced from Strava (sync pipeline)",
        "No",
        "— Backend pipeline only (result stored in MongoDB for RAG queries)",
        "Yes — stored in MongoDB as analysis field on activity embeddings via EmbeddingSync",
    ],

    # ── TOOL DESCRIPTIONS ──
    [
        "packages/sport-cycling/src/tools.ts",
        "8 tool descriptions (cycling tools)",
        "Lines 41-357 (descriptions embedded in tool())",
        "Tool metadata LLM uses to decide which cycling tools to call: calculate_zones, build_plan_skeleton, assess_feasibility, get_sample_week, validate_week_structure, build_taper_week, adjust_plan_for_missed_sessions, assess_model_fit.",
        "Injected as available tools in llm.generate({ tools: [...] })",
        "On every core agent chat() call — LLM decides which tool to invoke",
        "No",
        "— Core agent only (tool descriptions for LLM decision-making)",
        "No — tool metadata, not stored",
    ],
    [
        "packages/core/src/agent/tools.ts",
        "memory_read / memory_write tool descriptions",
        "Embedded in tool() definitions",
        "Tool descriptions for reading/writing memory sections. memory_write replaces entire section content.",
        "Injected as available tools in llm.generate({ tools: [...] })",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — writes to MEMORY.md on disk",
    ],
    [
        "packages/core/src/agent/strava-tools.ts",
        "Strava tool descriptions",
        "Embedded in tool() definitions",
        "Tool descriptions for strava_fetch_athlete, strava_fetch_activities, strava_fetch_activity, strava_search_history.",
        "Injected as available tools in llm.generate({ tools: [...] })",
        "On every core agent chat() call",
        "No",
        "— Core agent only",
        "No — tool metadata, not stored",
    ],
]

# Color fills for alternating section groups
fills = [
    PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid"),  # Backend
    PatternFill(start_color="E6F5FF", end_color="E6F5FF", fill_type="solid"),  # Sport-cycling
    PatternFill(start_color="F5FFF0", end_color="F5FFF0", fill_type="solid"),  # Skills
    PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid"),  # Core
    PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid"),  # Tools
]

# Section boundaries (0-indexed)
section_boundaries = [0, 5, 9, 14, 20, 25]  # start indices of each section

for row_idx, row_data in enumerate(data, 2):
    # determine fill based on section
    section_idx = -1
    for i, boundary in enumerate(section_boundaries):
        if row_idx - 2 >= boundary:
            section_idx = i

    fill = fills[section_idx] if 0 <= section_idx < len(fills) else None

    for col, value in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col, value=value)
        c.font = body_font
        c.alignment = wrap_alignment
        c.border = thin_border
        if fill:
            c.fill = fill

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:I{len(data) + 1}"

# Set row height for readability
for row in range(2, len(data) + 2):
    ws.row_dimensions[row].height = 50

ws.row_dimensions[1].height = 35

output_path = "C:\\Users\\piyus\\OneDrive\\Documents\\Workspaces\\cyclingCoach\\cycling-coach\\tools\\LLM_Prompts_Inventory.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Total prompts: {len(data)}")
